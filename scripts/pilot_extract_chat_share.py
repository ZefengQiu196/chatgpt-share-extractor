#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import ssl
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


SHARE_LINK_RE = re.compile(r"^https://chatgpt\.com/share/[A-Za-z0-9\-]+/?(?:\?.*)?$")
FLIGHT_PAYLOAD_RE = re.compile(
    r'window\.__reactRouterContext\.streamController\.enqueue\("(\[.*?\])\\n"\);',
    re.S,
)
CODE_BLOCK_RE = re.compile(r"```(?:[\w.+-]+)?\n(.*?)```", re.S)
ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


@dataclass
class TargetRecord:
    name_dot: str
    link: str


@dataclass
class StatusRecord:
    name_dot: str
    link: str
    status: str
    reason: str
    round_count: int


class ExtractionError(RuntimeError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pilot extractor for ChatGPT share links (c.2 + c.1994)."
    )
    parser.add_argument(
        "--metadata",
        default="AU25 CSE5544 LAB5 Metadata.xlsx",
        help="Path to metadata workbook.",
    )
    parser.add_argument(
        "--students",
        default="c.2,c.1994",
        help='Comma-separated Name.dot values to process, or "all".',
    )
    parser.add_argument(
        "--output-dir",
        default="pilot_results",
        help="Output directory for student xlsx files.",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=4,
        help="HTTP retries per share URL.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=35,
        help="HTTP timeout in seconds.",
    )
    parser.add_argument(
        "--status-file",
        default=None,
        help="Status xlsx path. Default: <output-dir>/_pilot_status.xlsx",
    )
    return parser.parse_args()


def xml_safe(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = ILLEGAL_XML_RE.sub("", text)
    return "".join(ch for ch in text if is_xml_char(ord(ch)))


def is_xml_char(codepoint: int) -> bool:
    return (
        codepoint in (0x9, 0xA, 0xD)
        or 0x20 <= codepoint <= 0xD7FF
        or 0xE000 <= codepoint <= 0xFFFD
        or 0x10000 <= codepoint <= 0x10FFFF
    )


def safe_sheet_title(raw: str) -> str:
    cleaned = re.sub(r"[\[\]\*\?/:\\]", "_", raw)
    cleaned = cleaned.strip() or "Sheet1"
    return cleaned[:31]


def read_targets(metadata_path: Path, wanted_students: set[str]) -> tuple[list[TargetRecord], list[StatusRecord]]:
    wb = load_workbook(metadata_path, data_only=True, read_only=True)
    ws = wb["metadata"] if "metadata" in wb.sheetnames else wb.active

    found: dict[str, TargetRecord] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name_dot = (row[1] or "").strip() if len(row) >= 2 and isinstance(row[1], str) else ""
        link = (row[2] or "").strip() if len(row) >= 3 and isinstance(row[2], str) else ""
        if name_dot in wanted_students and name_dot not in found:
            found[name_dot] = TargetRecord(name_dot=name_dot, link=link)
    wb.close()

    missing_status = []
    for student in sorted(wanted_students):
        if student not in found:
            missing_status.append(
                StatusRecord(
                    name_dot=student,
                    link="",
                    status="failed",
                    reason="student_not_found",
                    round_count=0,
                )
            )

    ordered_targets = [found[s] for s in sorted(found)]
    return ordered_targets, missing_status


def read_all_name_dots(metadata_path: Path) -> set[str]:
    wb = load_workbook(metadata_path, data_only=True, read_only=True)
    ws = wb["metadata"] if "metadata" in wb.sheetnames else wb.active

    students: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2 and isinstance(row[1], str):
            name_dot = row[1].strip()
            if name_dot:
                students.add(name_dot)
    wb.close()
    return students


def fetch_html(url: str, retries: int, timeout: int) -> str:
    last_exc: Exception | None = None
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        },
    )
    for attempt in range(1, retries + 1):
        try:
            with urllib.request.urlopen(request, timeout=timeout, context=ssl.create_default_context()) as resp:
                return resp.read().decode("utf-8", errors="replace")
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            if attempt < retries:
                time.sleep(0.8 * attempt)
    assert last_exc is not None
    raise last_exc


def decode_payload(html: str) -> list[Any]:
    match = FLIGHT_PAYLOAD_RE.search(html)
    if not match:
        raise ExtractionError("payload_not_found")

    escaped = match.group(1)
    try:
        # Decode JavaScript string escapes while preserving native Unicode characters.
        raw = json.loads(f'"{escaped}"')
    except json.JSONDecodeError as exc:
        raise ExtractionError("payload_string_decode_failed") from exc
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ExtractionError("payload_json_decode_failed") from exc
    if not isinstance(payload, list):
        raise ExtractionError("payload_not_list")
    return payload


def get_conversation(payload: list[Any]) -> dict[str, Any]:
    try:
        idx_mapping = payload.index("mapping")
        idx_current = payload.index("current_node")
        idx_conv_id = payload.index("conversation_id")
    except ValueError as exc:
        raise ExtractionError("conversation_keys_missing") from exc

    mapping_key = f"_{idx_mapping}"
    current_key = f"_{idx_current}"
    conv_id_key = f"_{idx_conv_id}"

    conv_index = None
    for i, value in enumerate(payload):
        if isinstance(value, dict) and mapping_key in value and current_key in value and conv_id_key in value:
            conv_index = i
            break
    if conv_index is None:
        raise ExtractionError("conversation_object_missing")

    @lru_cache(maxsize=None)
    def resolve_index(index: int) -> Any:
        value = payload[index]
        if isinstance(value, dict):
            output: dict[str, Any] = {}
            for key, ref in value.items():
                if isinstance(key, str) and key.startswith("_") and key[1:].isdigit():
                    key_index = int(key[1:])
                    real_key = payload[key_index] if 0 <= key_index < len(payload) else key
                else:
                    real_key = key
                output[str(real_key)] = resolve_ref(ref)
            return output
        if isinstance(value, list):
            return [resolve_ref(item) for item in value]
        return value

    def resolve_ref(ref: Any) -> Any:
        if ref == -5:
            return None
        if isinstance(ref, int) and 0 <= ref < len(payload):
            return resolve_index(ref)
        return ref

    conversation = resolve_index(conv_index)
    if not isinstance(conversation, dict):
        raise ExtractionError("conversation_resolve_failed")
    return conversation


def collect_main_chain(conversation: dict[str, Any]) -> list[dict[str, Any]]:
    mapping = conversation.get("mapping")
    current = conversation.get("current_node")
    if not isinstance(mapping, dict):
        raise ExtractionError("mapping_missing")
    if not isinstance(current, str):
        raise ExtractionError("current_node_missing")

    chain: list[dict[str, Any]] = []
    seen: set[str] = set()
    node_id = current

    while node_id and node_id not in seen and node_id in mapping:
        seen.add(node_id)
        node = mapping[node_id]
        if not isinstance(node, dict):
            break
        message = node.get("message")
        if isinstance(message, dict):
            author = message.get("author") if isinstance(message.get("author"), dict) else {}
            role = author.get("role")
            create_time = message.get("create_time")
            recipient = message.get("recipient")
            content = message.get("content") if isinstance(message.get("content"), dict) else {}
            content_type = content.get("content_type")
            parts = content.get("parts") if isinstance(content.get("parts"), list) else []
            text = "\n".join(part for part in parts if isinstance(part, str))
            metadata = message.get("metadata") if isinstance(message.get("metadata"), dict) else {}
            metadata_attachments = metadata.get("attachments") if isinstance(metadata.get("attachments"), list) else []
            direct_attachments = message.get("attachments") if isinstance(message.get("attachments"), list) else []
            attachments = metadata_attachments or direct_attachments
            upload_names: list[str] = []
            for attachment in attachments:
                if isinstance(attachment, dict):
                    name = attachment.get("name")
                    if isinstance(name, str):
                        cleaned = name.strip()
                        if cleaned:
                            upload_names.append(cleaned)

            thought_items_raw = content.get("thoughts") if isinstance(content.get("thoughts"), list) else []
            thought_items: list[dict[str, str]] = []
            for item in thought_items_raw:
                if isinstance(item, dict):
                    summary = item.get("summary")
                    detail = item.get("content")
                    thought_items.append(
                        {
                            "summary": summary.strip() if isinstance(summary, str) else "",
                            "content": detail.strip() if isinstance(detail, str) else "",
                        }
                    )

            reasoning_duration_sec: float | None = None
            finished_duration_sec = metadata.get("finished_duration_sec")
            if isinstance(finished_duration_sec, (int, float)):
                reasoning_duration_sec = float(finished_duration_sec)
            elif isinstance(finished_duration_sec, str):
                try:
                    reasoning_duration_sec = float(finished_duration_sec.strip())
                except ValueError:
                    reasoning_duration_sec = None

            reasoning_recap_text = ""
            recap_content = content.get("content")
            if isinstance(recap_content, str):
                reasoning_recap_text = recap_content.strip()

            chain.append(
                {
                    "role": role,
                    "recipient": recipient,
                    "content_type": content_type,
                    "create_time": create_time,
                    "text": text,
                    "uploads": upload_names,
                    "thought_items": thought_items,
                    "reasoning_duration_sec": reasoning_duration_sec,
                    "reasoning_recap_text": reasoning_recap_text,
                }
            )
        parent = node.get("parent")
        node_id = parent if isinstance(parent, str) else ""

    chain = [item for item in reversed(chain) if isinstance(item.get("create_time"), (int, float))]
    if not chain:
        raise ExtractionError("empty_chain")
    return chain


def extract_round_rows(chain: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rounds: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    for msg in chain:
        role = msg.get("role")
        text = (msg.get("text") or "").strip()
        if role == "user" and text:
            if current is not None:
                rounds.append(current)
            uploads = msg.get("uploads") if isinstance(msg.get("uploads"), list) else []
            prompt_upload = "; ".join(uploads) if uploads else "None"
            current = {
                "prompt": text,
                "prompt_upload": prompt_upload,
                "responses": [],
                "thought_blocks": [],
                "thought_times": [],
            }
            continue

        is_visible_assistant = (
            role == "assistant"
            and msg.get("recipient") == "all"
            and msg.get("content_type") in ("text", "thoughts", "reasoning_recap")
        )
        if current is None or not is_visible_assistant:
            continue

        content_type = msg.get("content_type")
        if content_type == "text" and text:
            current["responses"].append(text)
            continue

        if content_type == "thoughts":
            thought_items = msg.get("thought_items") if isinstance(msg.get("thought_items"), list) else []
            for item in thought_items:
                if not isinstance(item, dict):
                    continue
                summary = (item.get("summary") or "").strip()
                detail = (item.get("content") or "").strip()
                if summary and detail:
                    block = f"{summary}\n{detail}"
                else:
                    block = summary or detail
                if block:
                    current["thought_blocks"].append(block)
            continue

        if content_type == "reasoning_recap":
            duration_text = format_duration_text(
                msg.get("reasoning_duration_sec"),
                msg.get("reasoning_recap_text"),
            )
            if duration_text:
                current["thought_times"].append(duration_text)

    if current is not None:
        rounds.append(current)

    rows: list[dict[str, Any]] = []
    for idx, item in enumerate(rounds, start=1):
        response = "\n\n".join(item["responses"]).strip()
        thought_times = item.get("thought_times") if isinstance(item.get("thought_times"), list) else []
        thought_blocks = item.get("thought_blocks") if isinstance(item.get("thought_blocks"), list) else []
        thought_time_value = "; ".join(thought_times) if thought_times else "None"
        thought_value = build_thought_text(thought_blocks, thought_times)
        rows.append(
            {
                "Round": idx,
                "Prompt": item["prompt"],
                "Prompt_upload": item["prompt_upload"],
                "ChatGPT's thought time": thought_time_value,
                "ChatGPT's thought": thought_value,
                "ChatGPT's response": response,
                "ChatGPT's response code": extract_code_blocks(response),
            }
        )
    return rows


def format_duration_text(duration_seconds: Any, recap_text: Any) -> str:
    if isinstance(duration_seconds, (int, float)):
        value = float(duration_seconds)
        if value.is_integer():
            return f"{int(value)}s"
        display = f"{value:.2f}".rstrip("0").rstrip(".")
        return f"{display}s"

    if isinstance(recap_text, str):
        text = recap_text.strip()
        if not text:
            return ""
        match = re.search(r"thought\s+for\s+([0-9]+(?:\.[0-9]+)?)\s*(?:s|seconds?)", text, re.I)
        if match:
            num = float(match.group(1))
            if num.is_integer():
                return f"{int(num)}s"
            display = f"{num:.2f}".rstrip("0").rstrip(".")
            return f"{display}s"
    return ""


def build_thought_text(thought_blocks: list[Any], thought_times: list[Any]) -> str:
    clean_blocks = [str(block).strip() for block in thought_blocks if str(block).strip()]
    clean_times = [str(value).strip() for value in thought_times if str(value).strip()]

    if not clean_blocks and not clean_times:
        return "None"

    segments: list[str] = []
    for idx, block in enumerate(clean_blocks):
        segment = block
        if idx < len(clean_times):
            segment = f"{segment}\n\nThought for {clean_times[idx]}\nDone"
        segments.append(segment)

    if len(clean_times) > len(clean_blocks):
        for idx in range(len(clean_blocks), len(clean_times)):
            segments.append(f"Thought for {clean_times[idx]}\nDone")

    return "\n\n".join(segments) if segments else "None"


def extract_code_blocks(response_text: str) -> str:
    if not response_text:
        return ""
    blocks = [match.group(1).strip("\n") for match in CODE_BLOCK_RE.finditer(response_text)]
    blocks = [block for block in blocks if block.strip()]
    return "\n\n".join(blocks)


def write_student_workbook(path: Path, name_dot: str, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(name_dot)
    headers = [
        "Round",
        "Prompt",
        "Prompt_upload",
        "ChatGPT's thought time",
        "ChatGPT's thought",
        "ChatGPT's response",
        "ChatGPT's response code",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["Round"],
                xml_safe(row["Prompt"]),
                xml_safe(row["Prompt_upload"]),
                xml_safe(row["ChatGPT's thought time"]),
                xml_safe(row["ChatGPT's thought"]),
                xml_safe(row["ChatGPT's response"]),
                xml_safe(row["ChatGPT's response code"]),
            ]
        )
    wb.save(path)
    wb.close()


def write_status_workbook(path: Path, records: list[StatusRecord]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "status"
    ws.append(["Name.dot", "Link", "Status", "Reason", "Round_count"])
    for record in records:
        ws.append(
            [
                xml_safe(record.name_dot),
                xml_safe(record.link),
                xml_safe(record.status),
                xml_safe(record.reason),
                record.round_count,
            ]
        )
    wb.save(path)
    wb.close()


def classify_exception(exc: Exception) -> str:
    if isinstance(exc, urllib.error.HTTPError):
        return f"http_{exc.code}"
    if isinstance(exc, urllib.error.URLError):
        reason = getattr(exc, "reason", None)
        if reason is not None:
            return f"url_error:{reason}"
        return "url_error"
    if isinstance(exc, ExtractionError):
        return str(exc)
    return f"{exc.__class__.__name__}:{exc}"


def run() -> int:
    args = parse_args()
    metadata_path = Path(args.metadata)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    status_path = Path(args.status_file) if args.status_file else output_dir / "_pilot_status.xlsx"

    students_arg = args.students.strip().lower()
    if students_arg == "all":
        students = read_all_name_dots(metadata_path)
    else:
        students = {item.strip() for item in args.students.split(",") if item.strip()}
    targets, status_records = read_targets(metadata_path, students)

    print(f"Targets found: {len(targets)} / requested: {len(students)}")

    for target in targets:
        if not SHARE_LINK_RE.match(target.link):
            status_records.append(
                StatusRecord(
                    name_dot=target.name_dot,
                    link=target.link,
                    status="failed",
                    reason="invalid_share_link",
                    round_count=0,
                )
            )
            print(f"[FAILED] {target.name_dot}: invalid_share_link")
            continue

        try:
            html = fetch_html(target.link, retries=args.retries, timeout=args.timeout)
            payload = decode_payload(html)
            conversation = get_conversation(payload)
            chain = collect_main_chain(conversation)
            rows = extract_round_rows(chain)
            out_path = output_dir / f"{target.name_dot}.xlsx"
            write_student_workbook(out_path, target.name_dot, rows)
            status_records.append(
                StatusRecord(
                    name_dot=target.name_dot,
                    link=target.link,
                    status="success",
                    reason="",
                    round_count=len(rows),
                )
            )
            print(f"[SUCCESS] {target.name_dot}: rounds={len(rows)} -> {out_path}")
        except Exception as exc:  # noqa: BLE001
            reason = classify_exception(exc)
            status_records.append(
                StatusRecord(
                    name_dot=target.name_dot,
                    link=target.link,
                    status="failed",
                    reason=reason,
                    round_count=0,
                )
            )
            print(f"[FAILED] {target.name_dot}: {reason}")

    status_records.sort(key=lambda record: record.name_dot)
    write_status_workbook(status_path, status_records)
    print(f"Status file written: {status_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
